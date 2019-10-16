from django.shortcuts import render
from django.core.files.storage import FileSystemStorage
# from django.conf import settings
from upload.models import GeoFiles
# from upload.forms import GeoFilesForm
# from django.http import HttpResponseRedirect


# Create your views here.
def home(request):
    # documents = GeoFiles.objects.all()
    return render(request, 'home.html')


def upload_file(request):
    if request.method == 'POST' and request.FILES['address_file']:
        address_file = request.FILES['address_file']
        fs = FileSystemStorage()
        filename = fs.save(address_file.name, address_file)
        uploaded_file_url = fs.url(filename)

        data = find_location(filename)
        if data:
            return render(request, 'upload_file.html', {
                'uploaded_file_url': uploaded_file_url
            })
        else:
            return render(request, 'upload_file.html', {
                'error': "Something went wrong."
            })
    return render(request, 'upload_file.html')


def find_location(filename):
    print(filename)
    import os
    from openpyxl import load_workbook
    import googlemaps
    from geopy.geocoders import Nominatim

    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    file_path = os.path.join(BASE_DIR, 'media\\' + filename)
    wb = load_workbook(file_path)
    ws = wb.get_active_sheet()
    # min_row=1, max_col=3, max_row=2
    try:
        for index, row in enumerate(ws.rows, start=1):
            temp_len = len(row)
            user_agent = user_agent_string()
            geolocator = Nominatim(user_agent=user_agent)
            location = geolocator.geocode(row[0].value)
            ws.cell(row=index, column=temp_len + 1).value = location.latitude
            ws.cell(row=index, column=temp_len + 2).value = location.longitude
        wb.save(file_path)
        return True
    except Exception as e:
        return False


def user_agent_string():
    from random_user_agent.user_agent import UserAgent
    from random_user_agent.params import SoftwareName, OperatingSystem

    # you can also import SoftwareEngine, HardwareType, SoftwareType, Popularity from random_user_agent.params
    # you can also set number of user agents required by providing `limit` as parameter

    software_names = [SoftwareName.CHROME.value]
    operating_systems = [OperatingSystem.WINDOWS.value, OperatingSystem.LINUX.value]

    user_agent_rotator = UserAgent(software_names=software_names, operating_systems=operating_systems, limit=100)

    # # Get list of user agents.
    # user_agents = user_agent_rotator.get_user_agents()

    # Get Random User Agent String.
    user_agent = user_agent_rotator.get_random_user_agent()

    return user_agent
